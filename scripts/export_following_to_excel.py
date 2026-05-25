#!/usr/bin/env python3
"""
导出 CRM 中状态为「跟进中」的线索为 Excel 表格。
用法（项目根目录）：python scripts/export_following_to_excel.py
输出文件：data/leads_following_YYYYMMDD_HHMM.xlsx
"""
import sys
from pathlib import Path
from datetime import datetime

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv
load_dotenv(ROOT / ".env")


def _budget_columns(lead):
    """从 lead 解析出 (固定薪水, 时薪) 两列。优先用 extra，否则解析 budget_signal/salary_raw。"""
    extra = lead.get("extra") or {}
    fixed = (extra.get("fixed_budget") or "").strip()
    hourly = (extra.get("hourly_range") or "").strip()
    if fixed or hourly:
        return fixed, hourly
    raw = (lead.get("budget_signal") or lead.get("salary_raw") or "").strip()
    if not raw:
        return "", ""
    # 解析 "固定 1000 | 时薪 30-50" 或单一一种
    for part in raw.split("|"):
        part = part.strip()
        if part.startswith("固定"):
            fixed = part.replace("固定", "").strip()
        elif part.startswith("时薪"):
            hourly = part.replace("时薪", "").strip()
    return fixed, hourly


def _date_only(s):
    """把 ISO 时间串只保留日期部分，如 2026-03-02T05:40:59.270Z -> 2026-03-02。"""
    if not s:
        return ""
    s = str(s).strip()
    if "T" in s:
        return s.split("T")[0]
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return s


def main():
    from src.storage.store import list_leads, get_follow_up

    # 获取所有「跟进中」线索（分页取满）
    limit = 2000
    all_leads = []
    offset = 0
    while True:
        leads, total = list_leads(status_filter="following", limit=limit, offset=offset)
        all_leads.extend(leads)
        if len(leads) < limit or len(all_leads) >= total:
            break
        offset += limit

    if not all_leads:
        print("没有状态为「跟进中」的线索")
        return

    # 为每条补充跟进备注与更新时间
    for lead in all_leads:
        fu = get_follow_up(lead["id"])
        lead["_notes"] = (fu["notes"] or "") if fu else ""
        lead["_updated_at"] = (fu["updated_at"] or "") if fu else ""
        lead["_core_summary"] = (lead.get("extra") or {}).get("core_summary") or ""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("请先安装 openpyxl：pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "跟进中"

    # 表头
    headers = [
        "ID", "平台", "标题", "发布方", "固定薪水", "时薪", "发布时间", "抓取时间",
        "核心描述", "跟进备注", "备注更新时间", "直达链接",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)

    for row_idx, lead in enumerate(all_leads, 2):
        fixed_budget, hourly_budget = _budget_columns(lead)
        ws.cell(row=row_idx, column=1, value=lead.get("id") or "")
        ws.cell(row=row_idx, column=2, value=lead.get("platform") or "")
        ws.cell(row=row_idx, column=3, value=lead.get("title") or "")
        ws.cell(row=row_idx, column=4, value=lead.get("publisher") or "")
        ws.cell(row=row_idx, column=5, value=fixed_budget)
        ws.cell(row=row_idx, column=6, value=hourly_budget)
        ws.cell(row=row_idx, column=7, value=_date_only(lead.get("published_at")))
        ws.cell(row=row_idx, column=8, value=lead.get("crawled_at") or "")
        ws.cell(row=row_idx, column=9, value=lead.get("_core_summary") or "")
        ws.cell(row=row_idx, column=10, value=lead.get("_notes") or "")
        ws.cell(row=row_idx, column=11, value=lead.get("_updated_at") or "")
        ws.cell(row=row_idx, column=12, value=lead.get("source_url") or "")

    # 列宽
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 14   # 固定薪水
    ws.column_dimensions["F"].width = 14   # 时薪
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 50
    ws.column_dimensions["J"].width = 40
    ws.column_dimensions["K"].width = 22
    ws.column_dimensions["L"].width = 50

    out_dir = ROOT / "data"
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = out_dir / f"leads_following_{ts}.xlsx"
    wb.save(out_path)
    print(f"已导出 {len(all_leads)} 条「跟进中」线索到：{out_path}")


if __name__ == "__main__":
    main()
